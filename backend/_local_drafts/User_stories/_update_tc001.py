import openpyxl
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = openpyxl.load_workbook('ICS613_Group3_Test_Cases_Draft_V1.xlsx')
ws = wb['Manual Test Cases']

# TC-001 is row 2
# Columns: A=Test case ID, B=Related story, C=Scenario, D=Precondition,
#          E=Test data, F=Steps, G=Expected result, H=Actual result, I=Pass/fail, J=Automation Feasibility

new_steps = (
    "1. Admin logs in with admin@community.com. "
    "2. Admin navigates to the Invite Management page. "
    "3. Admin enters newbie@example.com and clicks Invite. "
    '4. Tester retrieves the invite token from the database: '
    "SELECT token FROM invites WHERE email = 'newbie@example.com' ORDER BY created_at DESC LIMIT 1; "
    "5. Tester opens the registration URL with the token (e.g., /register?token=<retrieved-token>). "
    "6. Tester completes registration with email=newbie@example.com, password=SecurePass123!, display_name=Jane Smith."
)

new_expected = (
    "a unique invite token is generated and linked to that email.; "
    "an invite email is automatically sent to that address with the registration link containing the token.; "
    'the admin sees the invite in the invite list with status "sent".; '
    "the retrieved token from the database matches the one in the invites table.; "
    "the registration with the token succeeds, creating an account with status EMAIL_PENDING.; "
    "a verification email is sent.; "
    "the invite token is marked as used."
)

ws['F2'].value = new_steps
ws['G2'].value = new_expected
ws['J2'].value = 'Manual (requires DB read access for token retrieval)'

wb.save('ICS613_Group3_Test_Cases_Draft_V1.xlsx')
print('TC-001 updated successfully.')
